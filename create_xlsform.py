"""
Generate XLSForm for Survey123: Farq_pilot_2-2-26
Full QA validation constraints embedded per the 21-error prevention spec.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()

# ============ SURVEY sheet ============
ws = wb.active
ws.title = 'survey'

headers = ['type', 'name', 'label', 'label::arabic', 'required', 'appearance',
           'calculation', 'relevant', 'constraint', 'constraint_message',
           'default', 'hint', 'hint::arabic', 'bind::esri:fieldType',
           'bind::esri:fieldLength']
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')

# Category relevance conditions
REST_CAFE = ("selected(${category}, 'Restaurants') or selected(${category}, 'Coffee Shops') "
             "or selected(${category}, 'Food and Beverages') or selected(${category}, 'Hospitality') "
             "or selected(${category}, 'Bakery') or selected(${category}, 'Fast Food')")
ATTRACTION = ("selected(${category}, 'Entertainment and Recreation') or selected(${category}, 'Sports')")
MOSQUE = "selected(${category}, 'Mosques')"

# Constraint: Arabic only (no A-Z)
AR_ONLY = "not(regex(., '[A-Za-z]'))"
AR_ONLY_MSG = "Arabic characters only / أحرف عربية فقط"

# Constraint: English only (no Arabic)
EN_ONLY = "not(regex(., '[\\u0600-\\u06FF]'))"
EN_ONLY_MSG = "English characters only / أحرف إنجليزية فقط"

# Constraint: not just a category word
AR_NOT_CATEGORY_ONLY = "string-length(.) > 3 and not(. = 'مطعم' or . = 'صيدلية' or . = 'مقهى' or . = 'مغسلة' or . = 'بقالة' or . = 'مسجد' or . = 'مدرسة' or . = 'مستشفى' or . = 'فندق')"
EN_NOT_CATEGORY_ONLY = "string-length(.) > 3 and not(. = 'Restaurant' or . = 'Pharmacy' or . = 'Cafe' or . = 'Laundry' or . = 'Grocery' or . = 'Mosque' or . = 'School' or . = 'Hospital' or . = 'Hotel')"

# Phone constraint: Saudi format
PHONE_CONSTRAINT = ("regex(., '^(\\+9665\\d{8}|05\\d{8}|\\+9661[1-9]\\d{6,7}|01[1-9]\\d{6,7}|800\\d{7}|900\\d{7}|920\\d{6}|911)$') "
                    "or . = 'UNAVAILABLE'")
PHONE_MSG = "Must be Saudi format: +9665XXXXXXXX, 05XXXXXXXX, 800/900/920..., or UNAVAILABLE"

# Email constraint
EMAIL_CONSTRAINT = "regex(., '^[^@]+@[^@]+\\.[^@]+$') or . = 'UNAVAILABLE'"
EMAIL_MSG = "Must be valid email or UNAVAILABLE"

# Website constraint: no Google Maps
WEBSITE_CONSTRAINT = "not(contains(., 'google.com/maps')) and not(contains(., 'goo.gl/maps')) and (regex(., '^https?://') or . = 'UNAVAILABLE')"
WEBSITE_MSG = "Must be a valid URL (not Google Maps). Use UNAVAILABLE if none."

# Commercial License: 10-11 digits
LICENSE_CONSTRAINT = "regex(., '^\\d{10,11}$') or . = 'UNAVAILABLE'"
LICENSE_MSG = "Must be 10-11 digits or UNAVAILABLE"

# Building number: 4 digits
BUILDING_CONSTRAINT = "regex(., '^\\d{4}$') or . = 'UNAVAILABLE'"
BUILDING_MSG = "Must be 4 digits or UNAVAILABLE"

def r(type_, name, label, label_ar, required='', appearance='', calc='', relevant='',
      constraint='', constraint_msg='', default='', hint='', hint_ar='', field_type='', field_len=''):
    """Helper to build a row with all columns."""
    return [type_, name, label, label_ar, required, appearance, calc, relevant,
            constraint, constraint_msg, default, hint, hint_ar, field_type, field_len]

rows = [
    # ======== IDENTITY ========
    r('begin group', 'identity', 'Identity', 'الهوية', '', 'field-list'),

    r('text', 'poi_name_ar', 'Name (Arabic) *', 'الاسم بالعربي *', 'yes', '',
      '', '', f'{AR_ONLY} and {AR_NOT_CATEGORY_ONLY} and string-length(.) >= 2',
      f'{AR_ONLY_MSG}. Must be 2+ chars, not just a category word.',
      '', 'الاسم الرسمي كما يظهر على اللوحة', 'Official name as on signboard',
      'esriFieldTypeString', '500'),

    r('text', 'poi_name_en', 'Name (English) *', 'الاسم بالإنجليزي *', 'yes', '',
      '', '', f'{EN_ONLY} and {EN_NOT_CATEGORY_ONLY} and string-length(.) >= 2',
      f'{EN_ONLY_MSG}. Must be 2+ chars, not just a category word.',
      '', 'Official name in English', 'الاسم الرسمي بالإنجليزي',
      'esriFieldTypeString', '500'),

    r('text', 'legal_name', 'Legal Name *', 'الاسم القانوني *', 'yes', '',
      '', '', 'string-length(.) >= 3',
      'Must be 3+ characters (as per commercial license)',
      '', 'As per commercial license (مؤسسة/شركة/LLC/Est.)', 'حسب الرخصة التجارية',
      'esriFieldTypeString', '500'),

    r('select_one category_list', 'category', 'Category *', 'التصنيف *', 'yes', 'autocomplete',
      '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('select_one subcategory_list', 'secondary_category', 'Secondary Category *', 'التصنيف الفرعي *',
      'yes', 'autocomplete',
      '', '', '', '', '', 'Must be consistent with Category', 'يجب أن يتوافق مع التصنيف',
      'esriFieldTypeString', '200'),

    r('text', 'category_level_3', 'Category Level 3', 'المستوى الثالث', '', '',
      '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('select_one company_status_list', 'company_status', 'Company Status *', 'حالة المنشأة *',
      'yes', '', '', '', '', '', 'Open', '', '', 'esriFieldTypeString', '50'),

    r('end group', 'identity', '', ''),

    # ======== OPERATIONS ========
    r('begin group', 'operations', 'Working Hours', 'ساعات العمل', '', 'field-list'),

    r('select_one workdays_list', 'working_days', 'Working Days *', 'أيام العمل *',
      'yes', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('select_one workhours_list', 'working_hours', 'Working Hours *', 'ساعات العمل *',
      'yes', '', '', '', '', '', '', 'Must match working days pattern', '',
      'esriFieldTypeString', '200'),

    r('select_one breaktime_list', 'break_time', 'Break Time *', 'وقت الاستراحة *',
      'yes', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '100'),

    r('select_one holidays_list', 'holidays', 'Holidays *', 'الإجازات *',
      'yes', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '100'),

    r('end group', 'operations', '', ''),

    # ======== LOCATION ========
    r('begin group', 'location_grp', 'Location', 'الموقع', '', 'field-list'),

    r('geopoint', 'poi_location', 'Coordinates (X/Y) * - Tap map', 'الإحداثيات * - اضغط الخريطة',
      'yes', 'maps', '', '', '', '', '', 'Must be in Riyadh area', 'يجب أن يكون في منطقة الرياض'),

    r('decimal', 'latitude', 'Latitude', 'خط العرض', '', 'hidden',
      'pulldata("@geopoint", ${poi_location}, "y")', '',
      '. >= 24.0 and . <= 25.5',
      'Latitude must be 24.0-25.5 (Riyadh area)', '', '', '', 'esriFieldTypeDouble', ''),

    r('decimal', 'longitude', 'Longitude', 'خط الطول', '', 'hidden',
      'pulldata("@geopoint", ${poi_location}, "x")', '',
      '. >= 46.0 and . <= 47.5',
      'Longitude must be 46.0-47.5 (Riyadh area)', '', '', '', 'esriFieldTypeDouble', ''),

    r('select_one entrance_list', 'entrance_location', 'Entrance Location *', 'موقع المدخل *',
      'yes', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '50'),

    r('text', 'building_number', 'Building Number', 'رقم المبنى', '', '',
      '', '', BUILDING_CONSTRAINT, BUILDING_MSG,
      '', '4 digits (e.g. 7311)', '4 أرقام', 'esriFieldTypeString', '50'),

    r('select_one floor_list', 'floor_number', 'Floor Number', 'رقم الطابق',
      '', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '20'),

    r('text', 'district_ar', 'District (Arabic)', 'الحي (عربي)', '', '',
      '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('text', 'district_en', 'District (English)', 'الحي (إنجليزي)', '', '',
      '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('text', 'google_map_url', 'Google Map URL', 'رابط قوقل ماب', '', '',
      '', '', "regex(., '^https://(www\\.)?google\\.com/maps') or regex(., '^https://goo\\.gl/maps') or . = '' or . = 'UNAVAILABLE'",
      'Must be a Google Maps URL', '', '', '', 'esriFieldTypeString', '1000'),

    r('end group', 'location_grp', '', ''),

    # ======== MEDIA (Exterior/Interior mandatory, no duplicate URLs) ========
    r('begin group', 'media_grp', 'Photos & Media', 'الصور والوسائط', '', 'field-list'),

    r('note', 'note_media_warn', '⚠ Exterior and Interior photos must be DIFFERENT. Menu photo is for Restaurants/Cafes only.',
      '⚠ يجب أن تكون صورة الخارج مختلفة عن الداخل. صورة القائمة للمطاعم/المقاهي فقط.'),

    r('image', 'exterior_photo', 'Business Exterior Photo *', 'صورة خارجية للمنشأة *',
      'yes', '', '', '', '', '',
      '', 'Clear photo of the storefront/entrance', 'صورة واضحة للواجهة/المدخل'),

    r('image', 'interior_photo', 'Business Interior Photo *', 'صورة داخلية للمنشأة *',
      'yes', '', '', '', '', '',
      '', 'Photo inside the establishment (must differ from exterior)', 'صورة من الداخل (يجب أن تختلف عن الخارجية)'),

    r('image', 'menu_photo', 'Menu Photo', 'صورة القائمة', '', '',
      '', REST_CAFE, '', '',
      '', 'Photo of the food/drink menu', 'صورة قائمة الطعام/المشروبات'),

    r('text', 'video_url', 'Video URL', 'رابط الفيديو', '', '',
      '', '', "regex(., '\\.(mp4|mov)') or regex(., 'youtube\\.com|youtu\\.be') or . = '' or . = 'UNAVAILABLE'",
      'Must be mp4/mov file or YouTube link',
      '', 'YouTube or direct video link (.mp4/.mov)', '', 'esriFieldTypeString', '2000'),

    r('end group', 'media_grp', '', ''),

    # ======== CONTACT ========
    r('begin group', 'contact_grp', 'Contact Information', 'معلومات التواصل', '', 'field-list'),

    r('text', 'phone_number', 'Phone Number', 'رقم الهاتف', '', 'numbers',
      '', '', PHONE_CONSTRAINT, PHONE_MSG,
      '', '+9665XXXXXXXX or 05XXXXXXXX', '', 'esriFieldTypeString', '50'),

    r('text', 'email', 'Email', 'البريد الإلكتروني', '', '',
      '', '', EMAIL_CONSTRAINT, EMAIL_MSG,
      '', 'name@domain.com', '', 'esriFieldTypeString', '200'),

    r('text', 'website', 'Website', 'الموقع الإلكتروني', '', '',
      '', '', WEBSITE_CONSTRAINT, WEBSITE_MSG,
      '', 'https://www.example.com (NOT Google Maps)', 'ليس رابط قوقل ماب', 'esriFieldTypeString', '500'),

    r('text', 'social_media', 'Instagram / TikTok / X / Snapchat', 'انستقرام / تيكتوك / إكس / سناب',
      '', '',
      '', '', "not(regex(., '^\\+?\\d{10,}$')) and not(contains(., 'wa.me')) and not(contains(., 'whatsapp'))",
      'Must be social profile URL, NOT phone number or WhatsApp',
      '', '@username or profile URL (no WhatsApp/phone)', 'رابط الحساب فقط (ليس واتساب)', 'esriFieldTypeString', '500'),

    r('select_one language_list', 'language', 'Language', 'اللغة',
      '', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '200'),

    r('select_one cuisine_list', 'cuisine', 'Cuisine', 'نوع المطبخ',
      '', 'autocomplete', '', REST_CAFE, '', '',
      '', 'Restaurants/Cafes only', 'للمطاعم والمقاهي فقط', 'esriFieldTypeString', '200'),

    r('select_one payment_list', 'payment_methods', 'Accepted Payment Methods', 'طرق الدفع المقبولة',
      '', '', '', '', '', '', '', '', '', 'esriFieldTypeString', '500'),

    r('text', 'commercial_license', 'Commercial License Number', 'رقم الرخصة التجارية', '', '',
      '', '', LICENSE_CONSTRAINT, LICENSE_MSG,
      '', '10-11 digit number', 'رقم من 10-11 خانة', 'esriFieldTypeString', '200'),

    r('text', 'menu_barcode_url', 'Menu Barcode URL', 'رابط باركود القائمة', '', '',
      '', REST_CAFE, '', '',
      '', 'Restaurants/Cafes only', 'للمطاعم والمقاهي فقط', 'esriFieldTypeString', '1000'),

    r('end group', 'contact_grp', '', ''),

    # ======== AMENITIES - ALL CATEGORIES ========
    r('begin group', 'amenities_all', 'General Amenities', 'المرافق العامة', '', 'field-list'),

    r('select_one yesno_list', 'wifi', 'WiFi', 'واي فاي'),
    r('select_one yesno_list', 'reservation', 'Reservation', 'حجز'),
    r('select_one yesno_list', 'pickup_point', 'Pickup Point Exists', 'نقطة استلام'),
    r('select_one yesno_list', 'children_area', 'Children Area', 'منطقة أطفال'),
    r('select_one yesno_list', 'valet_parking', 'Valet Parking', 'خدمة صف السيارات'),
    r('select_one yesno_list', 'music', 'Music', 'موسيقى'),
    r('select_one yesno_list', 'has_parking_lot', 'Has Parking Lot', 'موقف سيارات'),
    r('select_one yesno_list', 'wheelchair_accessible', 'Is Wheelchair Accessible', 'وصول كرسي متحرك'),

    r('end group', 'amenities_all', '', ''),

    # ======== AMENITIES - RESTAURANTS/CAFES ONLY ========
    r('begin group', 'amenities_fnb', 'Restaurant/Cafe Amenities', 'مرافق المطاعم والمقاهي',
      '', 'field-list', '', REST_CAFE),

    r('select_one yesno_list', 'menu', 'Menu Available', 'قائمة طعام متوفرة'),
    r('select_one yesno_list', 'drive_thru', 'Drive Thru', 'طلب من السيارة (درايف ثرو)'),
    r('select_one yesno_list', 'dine_in', 'Dine In', 'أكل في المطعم'),
    r('select_one yesno_list', 'only_delivery', 'Only Delivery', 'توصيل فقط'),
    r('select_one yesno_list', 'shisha', 'Shisha', 'شيشة'),
    r('select_one yesno_list', 'order_from_car', 'Order from Car', 'طلب من السيارة'),
    r('select_one yesno_list', 'live_sports', 'Live Sport Broadcasting', 'بث رياضي مباشر'),
    r('select_one yesno_list', 'family_seating', 'Has Family Seating', 'جلسات عائلية'),
    r('select_one yesno_list', 'large_groups', 'Large Groups Can Be Seated', 'استيعاب مجموعات كبيرة'),
    r('select_one yesno_list', 'waiting_area', 'Has a Waiting Area', 'منطقة انتظار'),
    r('select_one yesno_list', 'private_dining', 'Has Separate Rooms for Dining', 'غرف طعام خاصة'),
    r('select_one yesno_list', 'smoking_area', 'Has Smoking Area', 'منطقة تدخين'),
    r('select_one yesno_list', 'iftar_menu', 'Offers Iftar Menu', 'قائمة إفطار'),
    r('select_one yesno_list', 'open_suhoor', 'Is Open During Suhoor', 'مفتوح للسحور'),

    r('end group', 'amenities_fnb', '', ''),

    # ======== AMENITIES - MOSQUE ONLY ========
    r('begin group', 'amenities_mosque', 'Mosque Amenities', 'مرافق المساجد',
      '', 'field-list', '', MOSQUE),

    r('select_one yesno_list', 'women_prayer_room', 'Has Women-Only Prayer Room', 'مصلى نساء'),
    r('select_one yesno_list', 'iftar_tent', 'Provides Iftar Tent', 'خيمة إفطار'),

    r('end group', 'amenities_mosque', '', ''),

    # ======== AMENITIES - ATTRACTION/LANDMARK ONLY ========
    r('begin group', 'amenities_attraction', 'Attraction/Landmark', 'المعالم',
      '', 'field-list', '', ATTRACTION),

    r('select_one yesno_list', 'require_ticket', 'Require Ticket', 'يتطلب تذكرة'),
    r('select_one yesno_list', 'is_landmark', 'Is Landmark', 'معلم بارز'),
    r('select_one yesno_list', 'free_entry', 'Is Free Entry', 'دخول مجاني'),

    r('end group', 'amenities_attraction', '', ''),
]

for row in rows:
    ws.append(row)

# ============ CHOICES sheet ============
wc = wb.create_sheet('choices')
choice_headers = ['list_name', 'name', 'label', 'label::arabic']
wc.append(choice_headers)
for cell in wc[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='FFEEDD', end_color='FFEEDD', fill_type='solid')

# Yes/No/Unavailable
for val, en, ar in [('Yes','Yes','نعم'), ('No','No','لا'), ('UNAVAILABLE','Unavailable','غير متوفر'), ('UNAPPLICABLE','Not Applicable','غير قابل للتطبيق')]:
    wc.append(['yesno_list', val, en, ar])

# Company Status
for val, ar in [('Open','مفتوح'), ('Temporarily Closed','مغلق مؤقتاً'), ('Permanently Closed','مغلق نهائياً')]:
    wc.append(['company_status_list', val, val, ar])

# Floor
for val, en in [('UNAVAILABLE','Unavailable'), ('G','Ground Floor'), ('B1','Basement 1'), ('1','Floor 1'), ('2','Floor 2'), ('3','Floor 3'), ('4','Floor 4'), ('5','Floor 5')]:
    wc.append(['floor_list', val, en, en])

# Entrance Location
for val, ar in [('Front','أمامي'), ('Side','جانبي'), ('Back','خلفي'), ('Mall Entrance','مدخل المول'), ('Basement','الطابق السفلي'), ('Main Street','الشارع الرئيسي'), ('UNAVAILABLE','غير متوفر')]:
    wc.append(['entrance_list', val, val, ar])

# Categories
categories = [
    ('Accommodation', 'إقامة'), ('Automotive', 'سيارات'), ('Clinics & Labs', 'عيادات ومختبرات'),
    ('Coffee Shops', 'مقاهي'), ('Education', 'تعليم'), ('Entertainment and Recreation', 'ترفيه'),
    ('Finance', 'مالية'), ('Food and Beverages', 'أغذية ومشروبات'), ('Government', 'حكومي'),
    ('Healthcare', 'رعاية صحية'), ('Hospitality', 'ضيافة'), ('Mosques', 'مساجد'),
    ('Personal Care', 'عناية شخصية'), ('Public Services', 'خدمات عامة'), ('Real Estate', 'عقارات'),
    ('Restaurants', 'مطاعم'), ('Services', 'خدمات'), ('Shopping', 'تسوق'),
    ('Shopping & Distribution', 'تسوق وتوزيع'), ('Sports', 'رياضة'),
    ('Transportation', 'نقل'), ('Utilities', 'مرافق'),
]
for c, ar in categories:
    wc.append(['category_list', c, c, ar])

# Subcategories
subcategories = [
    'Serviced Apartment', 'Car Rental', 'Car Wash', 'Gas Station', 'Parking',
    'Clinic', 'Dental Clinic', 'Lab', 'Pharmacy', 'Optical',
    'Cafe', 'Coffee Shop', 'Juice Bar', 'Tea House',
    'School', 'University', 'Training Center', 'Nursery', 'Library',
    'Amusement Park', 'Cinema', 'Gym', 'Park', 'Sports Club', 'Stadium',
    'Bank', 'ATM', 'Exchange', 'Insurance',
    'Bakery', 'Catering', 'Fast Food', 'Fine Dining', 'Food Truck',
    'Hospital', 'Medical Center', 'Rehabilitation',
    'Hotel', 'Resort', 'Motel',
    'Mosque', 'Prayer Room',
    'Barber', 'Beauty Salon', 'Spa', 'Massage',
    'Restaurant', 'Buffet', 'Grill', 'Seafood Restaurant',
    'Tailor', 'Laundry', 'Printing', 'Travel Agency',
    'Mall', 'Supermarket', 'Bookstore', 'Electronics', 'Florist',
    'Bus Station', 'Airport', 'Metro Station', 'Taxi',
    'Support & Outsourcing Services', 'Contracting',
]
for s in subcategories:
    wc.append(['subcategory_list', s, s, s])

# Language
for val, ar in [('Arabic','عربي'), ('English','إنجليزي'), ('Arabic, English','عربي, إنجليزي'),
                ('Arabic, English, Urdu','عربي, إنجليزي, أردو'), ('Arabic, English, Filipino','عربي, إنجليزي, فلبيني'),
                ('Arabic, English, Hindi','عربي, إنجليزي, هندي'), ('Arabic, English, French','عربي, إنجليزي, فرنسي'),
                ('UNAVAILABLE','غير متوفر')]:
    wc.append(['language_list', val, val, ar])

# Cuisine
cuisines = [
    ('Fast Food','وجبات سريعة'), ('Mixed','متنوع'), ('Saudi / Arabic','سعودي / عربي'),
    ('International','عالمي'), ('Pizza / Italian','بيتزا / إيطالي'), ('Bakery / Pastry','مخبز / حلويات'),
    ('Lebanese','لبناني'), ('Middle Eastern','شرق أوسطي'), ('Indian','هندي'),
    ('Pakistani','باكستاني'), ('Chinese','صيني'), ('Japanese','ياباني'),
    ('Korean','كوري'), ('Thai','تايلندي'), ('Italian','إيطالي'),
    ('American','أمريكي'), ('Turkish','تركي'), ('Yemeni','يمني'),
    ('Egyptian','مصري'), ('Syrian','سوري'), ('Seafood','مأكولات بحرية'),
    ('Grills','مشويات'), ('Coffee / Cafe','قهوة / كافيه'),
    ('UNAVAILABLE','غير متوفر'), ('UNAPPLICABLE','غير قابل للتطبيق')
]
for c, ar in cuisines:
    wc.append(['cuisine_list', c, c, ar])

# Payment Methods
payments = [
    'Cash, Mada, Visa, Mastercard, Apple Pay', 'Cash, Mada, Visa, Mastercard',
    'Mada, Visa, Mastercard, Apple Pay', 'Cash, Mada, Visa, Apple Pay',
    'Cash, Mada, Mastercard, Apple Pay', 'Cash, Mada, Visa, Mastercard, Apple Pay, STC Pay',
    'Cash, Mada, Visa', 'Cash, Mada', 'Mada, Visa, Mastercard', 'Mada, Apple Pay',
    'Cash', 'Apple Pay', 'UNAVAILABLE'
]
for p in payments:
    wc.append(['payment_list', p, p, p])

# Working Days
workdays = ['Daily','Sun, Mon, Tue, Wed, Thu, Fri, Sat','Sun, Mon, Tue, Wed, Thu',
            'Sun, Mon, Tue, Wed, Thu, Sat','Sun, Mon, Tue, Wed, Thu, Fri',
            'Mon, Tue, Wed, Thu, Fri','UNAVAILABLE']
for w in workdays:
    wc.append(['workdays_list', w, w, w])

# Working Hours
workhours = ['06:00 - 22:00','06:00 - 00:00','07:00 - 23:00','08:00 - 22:00',
             '08:00 - 23:00','08:00 - 00:00','09:00 - 17:00','09:00 - 21:00',
             '09:00 - 22:00','09:00 - 23:00','09:00 - 00:00','10:00 - 22:00',
             '10:00 - 23:00','10:00 - 00:00','12:00 - 00:00','16:00 - 00:00',
             '16:00 - 02:00','Open 24 Hours','UNAVAILABLE']
for w in workhours:
    wc.append(['workhours_list', w, w, w])

# Break Time
breaktimes = ['No Break','12:00 - 13:00','12:30 - 13:30','13:00 - 14:00',
              '13:00 - 16:00','14:00 - 16:00','15:00 - 17:00','UNAVAILABLE','UNAPPLICABLE']
for b in breaktimes:
    wc.append(['breaktime_list', b, b, b])

# Holidays
holidays = ['No Holidays','Friday','Friday & Saturday','Saturday',
            'National Holidays Only','Ramadan Hours','UNAVAILABLE']
for h in holidays:
    wc.append(['holidays_list', h, h, h])

# ============ SETTINGS sheet ============
ws_settings = wb.create_sheet('settings')
ws_settings.append(['form_title', 'form_id', 'instance_name', 'style', 'version'])
for cell in ws_settings[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDFFDD', end_color='DDFFDD', fill_type='solid')
ws_settings.append(['Farq_pilot_2-2-26', 'Farq_pilot_2_2_26',
                     "concat(${poi_name_en}, ' - ', ${poi_name_ar})", 'theme-grid', '2026030703'])

# Auto-width columns
for sheet in [ws, wc, ws_settings]:
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 55)

output_path = r'C:\Users\abdul\media-review-app-1\Farq_pilot_2-2-26.xlsx'
wb.save(output_path)
print(f'XLSForm saved to: {output_path}')

# Summary
mandatory = sum(1 for row in rows if len(row) > 4 and row[4] == 'yes')
constrained = sum(1 for row in rows if len(row) > 8 and row[8])
relevant = sum(1 for row in rows if len(row) > 7 and row[7])
total_fields = sum(1 for row in rows if row[0] not in ('begin group', 'end group', 'note') and 'hidden' not in str(row[5] if len(row) > 5 else ''))
print(f'\nMandatory fields: {mandatory}')
print(f'Fields with validation constraints: {constrained}')
print(f'Fields with category relevance: {relevant}')
print(f'Visible fields: {total_fields}')
print(f'Choice lists: {wc.max_row - 1} options')

print(f'\n=== QA RULES EMBEDDED ===')
print(f'A) Name_AR: Arabic-only, no category-only words, min 2 chars')
print(f'B) Name_EN: English-only, no category-only words, min 2 chars')
print(f'C) Legal_Name: min 3 chars')
print(f'D) Media: exterior/interior required, menu/barcode F&B only, video mp4/mov/youtube')
print(f'E) Category: controlled vocabulary via select_one')
print(f'F) Company Status: Open/Temporarily Closed/Permanently Closed')
print(f'G) Coordinates: Riyadh bounds (lat 24-25.5, lon 46-47.5)')
print(f'H) Building: 4 digits, Floor: G/B1/1-5')
print(f'I) Phone: Saudi format, Email: valid@format, Website: no Google Maps, Social: no WhatsApp/phone')
print(f'J) Working Hours: from predefined list')
print(f'K) Booleans: Yes/No/UNAVAILABLE/UNAPPLICABLE, F&B fields hidden for non-F&B')
print(f'   Mosque fields shown only for Mosques')
print(f'   Attraction fields shown only for Entertainment/Sports')
